import openpyxl as op
from openpyxl.chart import BarChart,Reference

class DiagrammDraw():
    def diagramm(self,excel_file,sheet_name):
        df=op.load_workbook(excel_file)
        source=df[sheet_name]
        if 'Диаграмма' in df.sheetnames:
            del df['Диаграмма']
        chart_sheet=df.create_sheet('Диаграмма')
        chart=BarChart()
        chart.title='Расходы'
        chart.x_axis.title='Расходы'
        chart.y_axis.title='Рубли'
        max_row = source.max_row
        max_col = source.max_column
        data = Reference(source, min_col=1,max_col=max_col, min_row=1,max_row=max_row)
        chart.add_data(data,titles_from_data=True)
        chart_sheet.add_chart(chart, "A1")     
        df.save(excel_file)